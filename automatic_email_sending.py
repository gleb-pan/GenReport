import os
import time
import smtplib
import logging
import configparser
import pyodbc
import chardet
from datetime import datetime as dt, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter

filename = f"Alarms_{(dt.now() - timedelta(days=1)).strftime('%d.%m.%y')}"

# Configure logging globally
logging.basicConfig(filename='_internal\\app_log.log',
                    level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

print(
    '''
 _   _  ____  _____   _____ ______ _____ 
| \ | |/ __ \|  __ \ / ____|  ____/ ____|
|  \| | |  | | |__) | (___ | |__ | |     
| . ` | |  | |  _  / \___ \|  __|| |     
| |\  | |__| | | \ \ ____) | |___| |____ 
|_| \_|\____/|_|  \_\_____/|______\_____|
Automatic_email_sending_v1.2
    '''
)

time.sleep(1.5)


def log(message, *, e: bool = False, i: bool = False):
    if e:
        logging.error(message)
    if i:
        logging.info(message)


def _fetch_data_from_db(*, driver, server, database, username, password, query):
    connection_string = f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}'
    try:
        connection = pyodbc.connect(connection_string)
        cursor = connection.cursor()
        cursor.execute(query)

        columns = [column[0] for column in cursor.description]
        data = cursor.fetchall()

        cursor.close()
        connection.close()
        log(f'Fetched from database successfully ({server} -> {database})', i=True)
        print(f'Fetched from database successfully ({server} -> {database})')

        return columns, data

    except Exception as e:
        log(f'fetch_data_from_db(): {e}', e=True)
        print(f'fetch_data_from_db(): {e}')
        return None, None


def create_xlsx(db_data):
    if not os.path.exists('_internal/Data\\'):
        os.makedirs('_internal/Data')

    try:
        columns, data = db_data
        if not columns or not data:
            log("Failed to fetch data from database.", e=True)
            print("Failed to fetch data from database.")
            return None

        workbook = Workbook()
        worksheet = workbook.active

        # Define color fills
        fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_purple = PatternFill(start_color="ff00ff", end_color="ff00ff", fill_type="solid")

        # Define bold font style with increased size
        bold_font = Font(bold=True, size=16)

        # Append header row
        worksheet.append(columns)
        for cell in worksheet[1]:
            cell.font = bold_font  # Apply bold font to header row

        # Define column indexes for `alarm_class` and `log_action`
        alarm_class_idx = columns.index('alarm_class') + 1
        log_action_idx = columns.index('log_action') + 1

        # Append data rows and apply conditional formatting
        for row_idx, row in enumerate(data, start=2):
            row_list = list(row)  # Convert pyodbc.Row to list
            for col_idx, value in enumerate(row_list, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value

            # Apply conditional formatting
            alarm_class = row_list[alarm_class_idx - 1]
            log_action = row_list[log_action_idx - 1]

            if log_action == 'G':
                if alarm_class.lower() == 'high':
                    fill = fill_red
                elif alarm_class.lower() == 'med':
                    fill = fill_yellow
                elif alarm_class.lower() == 'low':
                    fill = fill_purple
                else:
                    continue  # Skip rows that do not meet any condition

                # Apply fill to the entire row
                for col in range(1, len(columns) + 1):
                    worksheet.cell(row=row_idx, column=col).fill = fill

        # Adjust column widths
        for col in range(1, len(columns) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in worksheet[column]:
                try:
                    if cell.value is not None:
                        value = str(cell.value)
                        if len(value) > max_length:
                            max_length = len(value)
                except Exception as e:
                    log(f"Error while measuring length of cell value: {e}", e=True)
                    print(f"Error while measuring length of cell value: {e}")
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
            # log(f"Column {column}: Width set to {adjusted_width}")

        # Add borders
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in range(1, len(data) + 2):  # +2 because we have header row
            for col in range(1, len(columns) + 1):
                worksheet.cell(row=row, column=col).border = border

        save_dir = os.path.join('.', '_internal/Data', f"{filename}.xlsx")
        workbook.save(save_dir)
        log(f'File successfully saved into directory ("{save_dir}")', i=True)
        print(f'File successfully saved into directory ("{save_dir}")')
        return save_dir
    except Exception as e:
        log(f'create_xlsx(): {e}', e=True)
        print(f'create_xlsx(): {e}')
        return None


def attach_file(*, message, path):
    if path and os.path.isfile(path):
        file_name = os.path.basename(path)
        file_size_mb = os.path.getsize(path) / (1024 ** 2)
        try:
            with open(path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={file_name}')
                message.attach(part)
            log(f'File "{file_name}" ({round(file_size_mb, 2)} MB) has been attached to the email.', i=True)
            print(f'File "{file_name}" ({round(file_size_mb, 2)} MB) has been attached to the email.')
        except Exception as e:
            log(f'attach_file(): {e}', e=True)
            print(f'attach_file(): {e}')


def send_daily_email(*, config, msg):
    try:
        log("Attempting to connect to the SMTP server...", i=True)
        server = smtplib.SMTP(config['Settings']['smtp_server'], config['Settings'].getint('port'), timeout=60)
        log("SMTP server connection established.", i=True)
        server.starttls()
        log("Started TLS session.", i=True)
        server.login(config['Credentials']['username'], config['Credentials']['password'])
        log("Logged into SMTP server.", i=True)

        print('Sending emails . . .')
        time.sleep(1.5)
        for recipient in config['Other']['to'].split(','):
            msg['To'] = recipient
            server.sendmail(config['Credentials']['username'], recipient, msg.as_string())
            log(f"Email sent successfully to {recipient}", i=True)
            print(f"Email sent successfully to {recipient}")
            time.sleep(1)

        server.quit()
        log("SMTP server connection closed.", i=True)
        print("SMTP server connection closed.")

    except Exception as e:
        log(f'send_daily_email(): {e}', e=True)


def main():
    try:
        yesterday = (dt.now() - timedelta(days=1)).strftime('%d.%m.%y')

        # LOADING DATA FROM CONFIG FILE
        # Detect encoding
        with open('_internal/config.ini', 'rb') as file:
            result = chardet.detect(file.read())
        encoding = result['encoding']

        # Read with detected encoding
        config = configparser.ConfigParser()
        with open('_internal/config.ini', 'r', encoding=encoding) as configfile:
            config.read_file(configfile)

        # ESTABLISHING THE CONNECTION WITH DATABASE
        db_data = _fetch_data_from_db(driver=config['db_conn']['driver']
                                   , server=config['db_conn']['server']
                                   , database=config['db_conn']['database']
                                   , username=config['db_conn']['username']
                                   , password=config['db_conn']['password']
                                   , query=config['db_conn']['QUERY']
                                   )
        # CREATING THE EMAIL
        msg = MIMEMultipart()
        msg['From'] = config['Credentials']['username']
        msg['Subject'] = f'BATYSPETROLEUM: Архив событий за {yesterday}'
        # Get current date and time
        now = dt.now()
        current_time = now.strftime("%d.%m.%Y %H:%M:%S")

        # Create the body of the email
        body = f'''
            <span style="color: red;"><i>{current_time}<br>Это письмо является автоматическим. Просим вас не отвечать на него.</i></span><br><br>
            Добрый день!<br><br>
            В приложении ежедневная выгрузка из системы SCADA.<br><br>
            <b>Best regards,</b><br>
            <b>NORSEC Team</b>
            '''

        msg.attach(MIMEText(body, 'html', 'utf-8'))

        # CREATING THE XLSX FILE
        xlsx_file = create_xlsx(db_data)
        if xlsx_file:
            attach_file(message=msg, path=xlsx_file)
            send_daily_email(config=config, msg=msg)
            log('============================================================================', i=True)
            print('Program finished.')
            time.sleep(1)
        else:
            log('Failed to create Excel file, no email sent.', e=True)
            print('Failed to create Excel file, no email sent.')

    except Exception as e:
        log(f'main(): {e}', e=True)
        print(f'Error: {e}')


if __name__ == '__main__':
    main()

