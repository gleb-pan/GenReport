import os
import smtplib
import logging
import configparser
from csv import reader, Sniffer
from datetime import datetime as dt, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from codecs import open as op
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

filename = f"Alarms_{(dt.now() - timedelta(days=1)).strftime('%d.%m.%y')}"

# Configure logging globally
logging.basicConfig(filename='app_log.log',
                    level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def log(message, *, e: bool = False, i: bool = False):
    if e:
        logging.error(message)
    if i:
        logging.info(message)


def read_csv_with_encoding(file_path, encoding):
    with op(file_path, 'r', encoding=encoding) as f:
        sample = f.read(1024)
        f.seek(0)
        dialect = Sniffer().sniff(sample)
        if dialect.delimiter not in [',', ';']:
            dialect.delimiter = ';'
        csv_data = [row for row in reader(f, dialect)]
        return csv_data


def create_xlsx(path):
    csv_data = []
    encodings = ['utf-8-sig', 'utf-16', 'utf-8']

    if not os.path.exists('.\\Data\\'):
        os.makedirs('.\\Data')

    try:
        for encoding in encodings:
            try:
                csv_data = read_csv_with_encoding(path, encoding)
                if csv_data:
                    break
            except Exception as e:
                log(f"Failed to read with encoding {encoding}: {e}", e=True)

        if not csv_data:
            log("Failed to read CSV with provided encodings.", e=True)
            return None

        row_count = len(csv_data)
        column_count = max(len(row) for row in csv_data)

        workbook = Workbook()
        worksheet = workbook.active

        for row in csv_data:
            worksheet.append(row)

        for col in range(1, column_count + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in worksheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for col in range(1, column_count + 1):
            worksheet.cell(row=1, column=col).fill = fill

        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in range(1, row_count + 1):
            for col in range(1, column_count + 1):
                worksheet.cell(row=row, column=col).border = border

        save_dir = os.path.join('.','Data', f"{filename}.xlsx")
        workbook.save(save_dir)
        log(f'File successfully saved into directory ("{save_dir}")', i=True)
        return save_dir
    except Exception as e:
        log(f'create_xlsx(): {e}', e=True)
        return None


def attach_file(*, message, path):
    if path and os.path.isfile(path):
        file_name = os.path.basename(path)
        file_size_mb = os.path.getsize(path) / (1024 ** 2)
        try:
            with open(path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={file_name}')
                message.attach(part)
            log(f'File "{file_name}" ({round(file_size_mb, 2)} MB) has been attached to the email.', i=True)
        except Exception as e:
            log(f'attach_file(): {e}', e=True)


def send_daily_email(*, config, msg):
    try:
        log("Attempting to connect to the SMTP server...", i=True)
        server = smtplib.SMTP(config['Settings']['smtp_server'], config['Settings'].getint('port'), timeout=60)
        log("SMTP server connection established.", i=True)
        server.starttls()
        log("Started TLS session.", i=True)
        server.login(config['Credentials']['username'], config['Credentials']['password'])
        log("Logged into SMTP server.", i=True)

        for recipient in config['Other']['to'].split(','):
            msg['To'] = recipient
            server.sendmail(config['Credentials']['username'], recipient, msg.as_string())
            log(f"Email sent successfully to {recipient}", i=True)

        server.quit()
        log("SMTP server connection closed.", i=True)

    except Exception as e:
        log(f'send_daily_email(): {e}', e=True)


def main():
    try:
        yesterday = (dt.now() - timedelta(days=1)).strftime('%d.%m.%y')

        config = configparser.ConfigParser()
        with open('config.ini', 'r', encoding='utf-8') as configfile:
            config.read_file(configfile)

        msg = MIMEMultipart()
        msg['From'] = config['Credentials']['username']
        msg['Subject'] = f'BATYSPETROLEUM: Архив событий за {yesterday}'

        body = config['Other']['email_body'].replace('\\n', '\n')
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        xlsx_file = create_xlsx(config['Other']['attachment_path'])
        if xlsx_file:
            attach_file(message=msg, path=xlsx_file)
            send_daily_email(config=config, msg=msg)
            log('============================================================================', i=True)
        else:
            log('Failed to create Excel file, no email sent.', e=True)

    except Exception as e:
        log(f'main(): {e}', e=True)


if __name__ == '__main__':
    main()
