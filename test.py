# from codecs import open as op
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Border, Side
# from openpyxl.utils import get_column_letter
# from csv import reader, Sniffer
# from datetime import datetime as dt
#
#
# # Get current time
# now = dt.now()
# today = now.strftime("%d.%m.%Y")
#
# # Path and filename
# filename = f'Архив_событий_{now.strftime("%d.%m.%Y")}'
#
#
# def read_csv_with_encoding(file_path, encoding):
#     with op(file_path, 'r', encoding=encoding) as f:
#         sample = f.read(1024)
#         f.seek(0)
#         dialect = Sniffer().sniff(sample)
#         if dialect.delimiter not in [',', ';']:
#             dialect.delimiter = ';'
#         csv_data = [row for row in reader(f, dialect)]
#         return csv_data
#
# # Try to read the CSV file with different encodings
# csv_data = []
# encodings = ['utf-8-sig', 'utf-16', 'utf-8']
# for encoding in encodings:
#     try:
#         csv_data = read_csv_with_encoding('Archive — копия.csv', encoding)
#         if csv_data:
#             break
#     except Exception as e:
#         print(f"Failed to read with encoding {encoding}: {e}")
#
# # Count rows and columns
# row_count = len(csv_data)
# column_count = max(len(row) for row in csv_data)
#
# # Write to Excel file
# workbook = Workbook()
# worksheet = workbook.active
#
# for row in csv_data:
#     worksheet.append(row)
#
# # 1. Automatically adjust the width of the columns
# for col in range(1, column_count + 1):
#     max_length = 0
#     column = get_column_letter(col)
#     for cell in worksheet[column]:
#         try:
#             if len(str(cell.value)) > max_length:
#                 max_length = len(cell.value)
#         except:
#             pass
#     adjusted_width = (max_length + 2)
#     worksheet.column_dimensions[column].width = adjusted_width
#
# # 2. Fill the very top row (only five columns) with blue color
# fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
# for col in range(1, column_count + 1):
#     worksheet.cell(row=1, column=col).fill = fill
#
# # 3. Apply "All borders" to the entire table
# thin = Side(border_style="thin", color="000000")
# border = Border(left=thin, right=thin, top=thin, bottom=thin)
#
# for row in range(1, row_count + 1):
#     for col in range(1, column_count + 1):
#         worksheet.cell(row=row, column=col).border = border
#
# # Save the workbook
# workbook.save(f"{filename}.xlsx")
#
# # Print total rows and columns
# print(f"Total rows: {row_count}")
# print(f"Total columns: {column_count}")


# import logging
#
# # Configure the logging
# logging.basicConfig(filename='logfile.txt',
#                     level=logging.WARNING,
#                     format='%(asctime)s - %(levelname)s - %(message)s')
#
# # Log some data
# logging.info("This is a log entry.")
#
# print("Data has been logged.")

import os
import smtplib
import logging
import configparser
from datetime import datetime as dt
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from codecs import open as op
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from csv import reader, Sniffer

filename = f'Alarms_{dt.now().strftime("%d.%m.%Y")}'

# Catching actions and errors
def log(message, *, e: bool = 0, i: bool = 0, file_name='app_log.log'):

    if e:
        logging.error(message)
    if i:
        logging.info(message)


def read_csv_with_encoding(file_path, encoding):
    with op(file_path, 'r', encoding=encoding) as f:
        sample = f.read(1024)
        f.seek(0)
        dialect = Sniffer().sniff(sample)
        if dialect.delimiter not in [',', ';']:
            dialect.delimiter = ';'
        csv_data = [row for row in reader(f, dialect)]
        return csv_data


def create_xlsx(path):
    # path - path for csv file
    csv_data = []
    encodings = ['utf-8-sig', 'utf-16', 'utf-8']

    # Checks if the Data folder exists
    if not os.path.exists('.\\Data\\'):
        os.makedirs('.\\Data')

    try:
        for encoding in encodings:
            try:
                csv_data = read_csv_with_encoding(path, encoding)
                if csv_data:
                    break
            except Exception as e:
                print(f"Failed to read with encoding {encoding}: {e}")

        # Count rows and columns
        row_count = len(csv_data)
        column_count = max(len(row) for row in csv_data)

        # Write to Excel file
        workbook = Workbook()
        worksheet = workbook.active

        for row in csv_data:
            worksheet.append(row)

        # 1. Automatically adjust the width of the columns
        for col in range(1, column_count + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in worksheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # 2. Fill the very top row (only five columns) with blue color
        fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for col in range(1, column_count + 1):
            worksheet.cell(row=1, column=col).fill = fill

        # 3. Apply "All borders" to the entire table
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in range(1, row_count + 1):
            for col in range(1, column_count + 1):
                worksheet.cell(row=row, column=col).border = border

        # Saving directory
        save_dir = f".\\Data\\{filename}.xlsx"

        # Save the workbook
        workbook.save(save_dir)
        log(f'File successfully saved into directory ("{save_dir}")', i=True)
        return save_dir

    except Exception as e:
        log(f'create_xlsx(): {e}', e=True)

    # Print total rows and columns
    # print(f"Total rows: {row_count}")
    # print(f"Total columns: {column_count}")


def attach_file(*, message, path):
    if path:
        file_name = os.path.basename(path)
        file_size_mb = (os.path.getsize(path))/(1024**2)
        # Attach the file
        try:
            with open(path, 'rb') as attachment:
                # Create a MIMEBase object
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                # Encode the payload using Base64
                encoders.encode_base64(part)
                # Add header to the part
                part.add_header('Content-Disposition', f'attachment; filename={file_name}')
                # Attach the part to the message
                message.attach(part)
            log(f'File "{file_name}" ({round(file_size_mb, 2)} MB) has been attached to the email.', i=True)
        except Exception as e:
            log(f'attach_file(): {e}', e=True)


def send_daily_email(*, config, msg):
    try:
        # Connect to the SMTP server
        server = smtplib.SMTP(config['Settings']['smtp_server'], config['Settings'].getint('port'))
        server.starttls()  # Secure the connection
        server.login(config['Credentials']['username'], config['Credentials']['password'])

        # Send the email to each recipient
        for recipient in config['Other']['to'].split(','):
            msg['To'] = recipient
            server.sendmail(config['Credentials']['username'], recipient, msg.as_string())
            log(f"Email sent successfully to {recipient}", i=True)

        # Quit the SMTP server
        server.quit()

    except Exception as e:
        log(f'send_daily_email(): {e}', e=True)


def main():
    try:
        # Current date
        today = dt.now().strftime('%d.%m.%y')

        # Taking data from config.ini file
        # config = configparser.ConfigParser()
        # config.read('config.ini')
        config = configparser.ConfigParser()
        with open('config.ini', 'r', encoding='utf-8') as configfile:
            config.read_file(configfile)

        # Create the email
        msg = MIMEMultipart()
        msg['From'] = config['Credentials']['username']
        msg['Subject'] = f'BATYSPETROLEUM: Архив событий за {today}'

        # Email body
        body = config['Other']['email_body'].replace('\\n', '\n')
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # Attaching the file and sending the email to the recipients
        xlsx_file = create_xlsx(config['Other']['attachment_path'])
        attach_file(message=msg, path=xlsx_file)
        send_daily_email(config=config, msg=msg)
        log('All emails are successfully sent.', i=True)
    except Exception as e:
        log(f'main(): {e}', e=True)


if __name__ == '__main__':
    #main()
    config = configparser.ConfigParser()
    with open('config.ini', 'r', encoding='utf-8') as configfile:
        config.read_file(configfile)
    print(config['Other']['to'].split(','))
