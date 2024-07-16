import glob
import logging
from codecs import open as op

#from shutil import move
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook
from csv import reader
from ctypes import windll
from datetime import datetime
from socket import gethostname
from openpyxl.styles import Alignment, Font, Border, Side
from os import getlogin

hostname = gethostname() # get PC-name using socket
username = getlogin() # get username using os

now = datetime.now() # Current time
today = now.strftime("%d.%m.%Y_%H-%M") # Filtered current time
final_path = f"C:\\Users\\{username}\\Desktop" # path where the file will be saved to
filename = f'Отчет_по_моточасам_{today}' # name of resulted .xlsx file


def check_name(f_path):
    global filename
    res = glob.glob(f"{f_path}\\{filename}*.xlsx")  # looking for the file by f_name
    if res:  # if there is any match
        sz = len(res)
        filename += f" ({sz})"


# def move_file(f_name, f_path):
#     res = glob.glob(f"{f_path}\\{f_name}*.xlsx") # looking for the file by f_name
#     if res: # if there is any match
#         sz = len(res)
#         f_name = f_name + f" ({sz})"
#     # if list is empty then there is no match found
#     move(f_name+".xlsx", f_path)


# EDIT THE FILE HERE
def adjust_and_save_sheet(*, A_width = 35, text):
    # INIT
    ft = Font(size=15, bold=True)
    all_borders = Border(left=Side(style='thin')
                         , right=Side(style='thin')
                         , top=Side(style='thin')
                         , bottom=Side(style='thin')
                         )
    #final_path + "\\" + filename + ".xlsx"
    ws = load_workbook(f"{final_path}\\{filename}.xlsx")
    ws.encoding = "windows-1252"
    sheet = ws.active

    # FILTERING

    # Inserting the timestamp to the very top(merged) cell
    sheet.column_dimensions['A'].width = A_width
    sheet.insert_rows(1) # inserting emtry row at the begining
    sheet.merge_cells('A1:D1') # merging cells
    cell = sheet['A1'] # selecting merged cell
    cell.font = ft # changing font
    cell.value = text
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Making A2:D2 bold
    ft = Font(size=11, bold=True)
    cells = sheet['A2:D2']
    for i in cells:
        for j in i:
            j.font = ft

    # Making borders for all cells
    cells = sheet['A1:D90']
    for i in cells:
        for j in i:
            j.border = all_borders

    # Changing cell format to 'number' for B3:D90
    cells = sheet['B3:D90']
    for i in cells:
        for j in i:
            if j.value is not None and j.value.isdigit():
                j.value = int(j.value)


    # SAVING
    ws.save(f"{final_path}\\{filename}.xlsx")


def generator(howmany):
    for x in howmany:
        yield x


def csv_to_xlsx(csv_filename):
    # Read CSV file
    csv_data = []
    try:
        encd = None
        with op(csv_filename, 'r', encoding=encd) as f:
            csv_data = [row for row in reader(f)]

    except:
        encd = "utf-16"
        with op(csv_filename, 'r', encoding=encd) as f:
            csv_data = [row for row in reader(f)]

    g = generator(csv_data)
    # Write to Excel file
    workbook = Workbook()
    worksheet = workbook.active
    for row in g:
        worksheet.append(row)
    workbook.save(f"{final_path}\\{filename}.xlsx")


def write_to_log(date_time, pc_name, result, file_name, error):
    logging.basicConfig(filename='C:\\Users\\Public\\Documents\\GenReport\\log.txt', filemode='a', format='%(message)s')
    delimiter = "==============================================" if (error is None) else "**********************************************"
    text = f"{delimiter}\nDate: {date_time}\nPC_name: {pc_name}\nResult: {result}\n" \
           f"Generated filename: {file_name}\n" \
           f"Error: {error}\n{delimiter}\n"
    logging.error(text) #if (error is None) else logging.error(text)


def gen_device_report(*, delay=1):
    # FIRST POP-UP
    windll.user32.MessageBoxW(0, "Нажмите 'OK', чтобы сохранить на рабочий стол", "Отчеты", 0)

    sleep(delay)  # time delay

    check_name(f_path=final_path)
    # reading the csv file
    csv_to_xlsx('C:\\Users\\Public\\Documents\\GenReport\\Data.csv')

    # EDITING EXCEL FILE AND MOVE FILE TO DESKTOP
    timestamp = f"Дата выгрузки: {now.strftime('%d.%m.%Y %H:%M')}"
    adjust_and_save_sheet(text=timestamp)

    # MOVE FILE TO DESKTOP
    # move_file(filename, final_path)  # moving the file using the os library

    windll.user32.MessageBoxW(0, f"Cохранено под именем: {filename}", "Отчеты", 0)
    # LOGGING
    write_to_log(date_time=now, pc_name=hostname, result='Success', file_name=filename, error=None)


# MAIN
if __name__ == '__main__':
    try:
        gen_device_report()
    except Exception as e:
        windll.user32.MessageBoxW(0, f"Ошибка: {e}", "Отчеты", 0)
        # LOGGING
        write_to_log(date_time=now, pc_name=hostname, result='Failed', file_name=filename, error=e)
        raise e
